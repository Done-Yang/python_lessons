from openpyxl import load_workbook
import mysql.connector

# Excel
my_workbook = load_workbook('Book1.xlsx')
sheet = my_workbook.active

values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)
    values.append(row)

db = mysql.connector.connect(
    host='localhost',
    port=3306,
    user='root',
    password='doneMySQL1582022',
    database='done_want_to_buy'
)

cursor = db.cursor()
sql = '''
    INSERT INTO products (title, price, is_necessary)
    VALUES (%s, %s, %s);
'''

cursor.executemany(sql, values)

db.commit()
print('add number of data: ' + str(cursor.rowcount) + ' row')
